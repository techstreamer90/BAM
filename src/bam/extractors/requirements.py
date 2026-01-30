"""
Requirements Extractor

Extracts design requirements from Excel (.xlsx) and CSV files.
Handles common column patterns and traceability relationships.
"""

import csv
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from .base import BaseExtractor, ExtractionResult
from .intermediate import IntermediateData, IntermediateNode, IntermediateEdge


logger = logging.getLogger("bam.extractors.requirements")


# Common column name patterns for auto-detection
# Note: Patterns use [-_\s]? to match hyphens, underscores, or spaces
COLUMN_PATTERNS = {
    # ID column
    "id": [
        r"^id$", r"^req(uirement)?[-_\s]?id$", r"^dr[-_\s]?id$",
        r"^item[-_\s]?id$", r"^number$", r"^#$", r"^key$"
    ],
    # Name/Title column
    "name": [
        r"^name$", r"^title$", r"^req(uirement)?[-_\s]?name$",
        r"^short[-_\s]?name$", r"^summary$"
    ],
    # Description column
    "description": [
        r"^desc(ription)?$", r"^text$", r"^body$", r"^content$",
        r"^requirement[-_\s]?text$", r"^details?$", r"^specification$", r"^spec$"
    ],
    # Status column
    "status": [
        r"^status$", r"^state$", r"^workflow[-_\s]?status$", r"^review[-_\s]?status$"
    ],
    # Priority column
    "priority": [
        r"^priority$", r"^prio$", r"^importance$", r"^weight$", r"^severity$"
    ],
    # Type/Category column
    "type": [
        r"^type$", r"^category$", r"^req[-_\s]?type$", r"^classification$",
        r"^level$", r"^kind$"
    ],
    # Owner/Assignee column
    "owner": [
        r"^owner$", r"^assignee$", r"^assigned[-_\s]?to$", r"^responsible$",
        r"^author$", r"^created[-_\s]?by$"
    ],
    # Traceability columns (can have multiple)
    "traces_to": [
        r"^traces?[-_\s]?to$", r"^parent$", r"^parent[-_\s]?id$", r"^derives[-_\s]?from$",
        r"^linked[-_\s]?to$", r"^related[-_\s]?to$", r"^satisfies$", r"^implements$",
        r"^upstream$", r"^source[-_\s]?req(uirement)?$"
    ],
    # Verification/Test links
    "verified_by": [
        r"^verified[-_\s]?by$", r"^test[-_\s]?(case)?[-_\s]?id$", r"^validation$",
        r"^vv[-_\s]?method$"
    ],
    # Version/Revision
    "version": [
        r"^version$", r"^rev(ision)?$", r"^baseline$"
    ],
    # Rationale
    "rationale": [
        r"^rationale$", r"^justification$", r"^reason$", r"^why$"
    ],
}


@dataclass
class RequirementItem:
    """Represents a requirement extracted from a spreadsheet."""
    id: str
    name: str
    description: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    req_type: Optional[str] = None
    owner: Optional[str] = None
    version: Optional[str] = None
    rationale: Optional[str] = None
    traces_to: List[str] = field(default_factory=list)
    verified_by: List[str] = field(default_factory=list)
    # Source tracking
    source_file: str = ""
    row_number: int = 0
    # Additional properties from unmapped columns
    extra_properties: Dict[str, Any] = field(default_factory=dict)


class RequirementsExtractor(BaseExtractor):
    """Extractor for requirements from Excel and CSV files."""

    EXTRACTOR_TYPE = "requirements"
    FILE_EXTENSIONS = ["xlsx", "csv"]

    def __init__(
        self,
        source_id: str,
        project_dir: Optional[str] = None,
        node_type: str = "Requirement",
        traces_edge_type: str = "TRACES_TO",
        verified_edge_type: str = "VERIFIED_BY",
    ):
        """Initialize the requirements extractor.

        Args:
            source_id: Identifier for this data source.
            project_dir: Optional project directory.
            node_type: Node type to use for requirements (default: "Requirement").
            traces_edge_type: Edge type for traceability links.
            verified_edge_type: Edge type for verification links.
        """
        super().__init__(source_id, project_dir)
        self.node_type = node_type
        self.traces_edge_type = traces_edge_type
        self.verified_edge_type = verified_edge_type

    def extract_file(self, file_path: Path) -> Dict[str, Any]:
        """Extract requirements from a single file.

        Args:
            file_path: Path to the Excel or CSV file.

        Returns:
            Dictionary with items, relationships, and metadata.
        """
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()

        if suffix == ".xlsx":
            return self._extract_excel(file_path)
        elif suffix == ".csv":
            return self._extract_csv(file_path)
        else:
            return {
                "items": [],
                "relationships": [],
                "warnings": [f"Unsupported file type: {suffix}"]
            }

    def _extract_excel(self, file_path: Path) -> Dict[str, Any]:
        """Extract from Excel file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return {
                "items": [],
                "relationships": [],
                "warnings": ["openpyxl not installed. Run: pip install openpyxl"]
            }

        items = []
        relationships = []
        warnings = []

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))

                if len(rows) < 2:
                    warnings.append(f"Sheet '{sheet_name}' has no data rows")
                    continue

                # First row is headers
                headers = [str(h).strip() if h else "" for h in rows[0]]
                column_map = self._detect_columns(headers)

                if not column_map.get("id"):
                    warnings.append(f"Sheet '{sheet_name}': No ID column detected")
                    continue

                # Process data rows
                for row_idx, row in enumerate(rows[1:], start=2):
                    req = self._parse_row(
                        row, headers, column_map,
                        str(file_path), row_idx, sheet_name
                    )
                    if req and req.id:
                        items.append(self._requirement_to_dict(req))
                        relationships.extend(self._extract_relationships(req))

            workbook.close()

        except Exception as e:
            warnings.append(f"Error reading Excel file: {e}")
            logger.exception("Error reading Excel file %s", file_path)

        return {
            "items": items,
            "relationships": relationships,
            "warnings": warnings,
            "metadata": {
                "file_path": str(file_path),
                "file_type": "excel",
            }
        }

    def _extract_csv(self, file_path: Path) -> Dict[str, Any]:
        """Extract from CSV file."""
        items = []
        relationships = []
        warnings = []

        try:
            # Detect encoding
            encoding = self._detect_encoding(file_path)

            with open(file_path, 'r', encoding=encoding, newline='') as f:
                # Detect delimiter
                sample = f.read(4096)
                f.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                except csv.Error:
                    dialect = csv.excel  # Default to comma-separated

                reader = csv.reader(f, dialect)
                rows = list(reader)

            if len(rows) < 2:
                return {
                    "items": [],
                    "relationships": [],
                    "warnings": ["CSV file has no data rows"]
                }

            # First row is headers
            headers = [str(h).strip() for h in rows[0]]
            column_map = self._detect_columns(headers)

            if not column_map.get("id"):
                return {
                    "items": [],
                    "relationships": [],
                    "warnings": ["No ID column detected in CSV"]
                }

            # Process data rows
            for row_idx, row in enumerate(rows[1:], start=2):
                # Pad row to match headers if needed
                while len(row) < len(headers):
                    row.append("")

                req = self._parse_row(
                    row, headers, column_map,
                    str(file_path), row_idx
                )
                if req and req.id:
                    items.append(self._requirement_to_dict(req))
                    relationships.extend(self._extract_relationships(req))

        except Exception as e:
            warnings.append(f"Error reading CSV file: {e}")
            logger.exception("Error reading CSV file %s", file_path)

        return {
            "items": items,
            "relationships": relationships,
            "warnings": warnings,
            "metadata": {
                "file_path": str(file_path),
                "file_type": "csv",
            }
        }

    def _detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding, defaulting to utf-8."""
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
            result = chardet.detect(raw)
            return result['encoding'] or 'utf-8'
        except ImportError:
            pass

        # Try common encodings
        for encoding in ['utf-8', 'utf-8-sig', 'cp1252', 'iso-8859-1']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1000)
                return encoding
            except UnicodeDecodeError:
                continue

        return 'utf-8'

    def _detect_columns(self, headers: List[str]) -> Dict[str, List[int]]:
        """Detect which columns map to which requirement fields.

        Returns a dict mapping field names to list of column indices.
        """
        column_map: Dict[str, List[int]] = {key: [] for key in COLUMN_PATTERNS}

        for idx, header in enumerate(headers):
            if not header:
                continue

            header_lower = header.lower().strip()

            for field_name, patterns in COLUMN_PATTERNS.items():
                for pattern in patterns:
                    if re.match(pattern, header_lower, re.IGNORECASE):
                        column_map[field_name].append(idx)
                        break

        # Log detected mappings
        detected = {k: v for k, v in column_map.items() if v}
        if detected:
            logger.info("Detected columns: %s", detected)

        return column_map

    def _parse_row(
        self,
        row: tuple,
        headers: List[str],
        column_map: Dict[str, List[int]],
        file_path: str,
        row_number: int,
        sheet_name: Optional[str] = None
    ) -> Optional[RequirementItem]:
        """Parse a single row into a RequirementItem."""

        def get_value(field: str) -> Optional[str]:
            """Get first non-empty value from mapped columns."""
            for idx in column_map.get(field, []):
                if idx < len(row) and row[idx]:
                    val = str(row[idx]).strip()
                    if val:
                        return val
            return None

        def get_list(field: str) -> List[str]:
            """Get list of values from mapped columns, splitting on common delimiters."""
            values = []
            for idx in column_map.get(field, []):
                if idx < len(row) and row[idx]:
                    cell_value = str(row[idx]).strip()
                    if cell_value:
                        # Split on common delimiters
                        parts = re.split(r'[,;|\n]+', cell_value)
                        for part in parts:
                            part = part.strip()
                            if part:
                                values.append(part)
            return values

        # Get required ID
        req_id = get_value("id")
        if not req_id:
            return None

        # Get name, falling back to description snippet if no name column
        req_name = get_value("name")
        if not req_name:
            desc = get_value("description")
            if desc:
                # Use first 50 chars of description as name
                req_name = desc[:50] + ("..." if len(desc) > 50 else "")
            else:
                req_name = req_id  # Fall back to ID

        # Build source location
        source_loc = file_path
        if sheet_name:
            source_loc = f"{file_path}:{sheet_name}:{row_number}"
        else:
            source_loc = f"{file_path}:{row_number}"

        # Collect extra properties from unmapped columns
        mapped_indices: Set[int] = set()
        for indices in column_map.values():
            mapped_indices.update(indices)

        extra_props = {}
        for idx, header in enumerate(headers):
            if idx not in mapped_indices and idx < len(row) and row[idx]:
                val = str(row[idx]).strip()
                if val and header:
                    extra_props[header] = val

        return RequirementItem(
            id=req_id,
            name=req_name,
            description=get_value("description"),
            status=get_value("status"),
            priority=get_value("priority"),
            req_type=get_value("type"),
            owner=get_value("owner"),
            version=get_value("version"),
            rationale=get_value("rationale"),
            traces_to=get_list("traces_to"),
            verified_by=get_list("verified_by"),
            source_file=file_path,
            row_number=row_number,
            extra_properties=extra_props,
        )

    def _requirement_to_dict(self, req: RequirementItem) -> Dict[str, Any]:
        """Convert RequirementItem to dictionary for ExtractionResult."""
        props = {}

        if req.status:
            props["status"] = req.status
        if req.priority:
            props["priority"] = req.priority
        if req.req_type:
            props["type"] = req.req_type
        if req.owner:
            props["owner"] = req.owner
        if req.version:
            props["version"] = req.version
        if req.rationale:
            props["rationale"] = req.rationale
        if req.traces_to:
            props["traces_to"] = req.traces_to
        if req.verified_by:
            props["verified_by"] = req.verified_by

        # Add extra properties
        props.update(req.extra_properties)

        source_loc = f"{req.source_file}:{req.row_number}"

        return {
            "id": f"req-{req.id}",
            "type": self.node_type,
            "name": req.name,
            "description": req.description,
            "file_path": req.source_file,
            "line_number": req.row_number,
            "properties": props,
        }

    def _extract_relationships(self, req: RequirementItem) -> List[Dict[str, Any]]:
        """Extract relationships from a requirement."""
        relationships = []

        # Traceability links
        for target_id in req.traces_to:
            relationships.append({
                "type": self.traces_edge_type,
                "source": req.id,
                "target": target_id.strip(),
                "source_type": "req",
                "target_type": "req",
            })

        # Verification links
        for target_id in req.verified_by:
            relationships.append({
                "type": self.verified_edge_type,
                "source": req.id,
                "target": target_id.strip(),
                "source_type": "req",
                "target_type": "test",
            })

        return relationships

    def to_intermediate(self, result: ExtractionResult) -> IntermediateData:
        """Convert extraction result to intermediate format.

        This is the preferred output format for Stage 2 ingestion.
        """
        intermediate = IntermediateData(
            source_id=result.source_id,
            source_type=self.EXTRACTOR_TYPE,
            extraction_metadata={
                "file_count": result.file_count,
                **result.metadata
            },
            warnings=result.warnings,
        )

        # Convert items to intermediate nodes
        for item in result.items:
            node = IntermediateNode(
                type=item["type"],
                external_id=item["id"],
                label=item["name"],
                properties=item.get("properties", {}),
                content=item.get("description"),
                source_location=f"{item['file_path']}:{item['line_number']}",
            )
            intermediate.add_node(node)

        # Convert relationships to intermediate edges
        for rel in result.relationships:
            edge = IntermediateEdge(
                type=rel["type"],
                source_external_id=f"{rel.get('source_type', 'req')}-{rel['source']}",
                target_external_id=f"{rel.get('target_type', 'req')}-{rel['target']}",
                properties={k: v for k, v in rel.items()
                           if k not in ("type", "source", "target", "source_type", "target_type")},
            )
            intermediate.add_edge(edge)

        return intermediate

    def generate_summary(self, result: ExtractionResult) -> Dict[str, Any]:
        """Generate a summary to help agents understand the requirements.

        Provides:
        - Counts by status, priority, type
        - Traceability statistics
        - Top-level (root) requirements that don't derive from others
        - Leaf requirements that aren't derived from
        """
        summary = {
            "total_requirements": len(result.items),
            "total_relationships": len(result.relationships),
            "by_status": {},
            "by_priority": {},
            "by_type": {},
            "files_processed": result.file_count,
            "top_level_requirements": [],  # Root reqs (don't trace to anything)
            "leaf_requirements": [],  # Not traced to by anything
            "orphan_requirements": [],  # No relationships at all
            "traceability": {
                "with_traces": 0,
                "with_verification": 0,
                "isolated": 0,
            }
        }

        # Track which requirements are traced to (have incoming edges)
        traced_to = set()
        for rel in result.relationships:
            if rel.get("type") == self.traces_edge_type:
                traced_to.add(rel.get("target"))

        for item in result.items:
            props = item.get("properties", {})

            # Count by status
            status = props.get("status", "unknown")
            summary["by_status"][status] = summary["by_status"].get(status, 0) + 1

            # Count by priority
            priority = props.get("priority", "unknown")
            summary["by_priority"][priority] = summary["by_priority"].get(priority, 0) + 1

            # Count by type
            req_type = props.get("type", "unknown")
            summary["by_type"][req_type] = summary["by_type"].get(req_type, 0) + 1

            req_id = item["id"].replace("req-", "")

            # Traceability stats
            traces = props.get("traces_to", [])
            verified = props.get("verified_by", [])

            if traces:
                summary["traceability"]["with_traces"] += 1
            if verified:
                summary["traceability"]["with_verification"] += 1

            # Top-level (root) requirements don't trace to anything
            if not traces:
                summary["top_level_requirements"].append(req_id)

            # Leaf requirements aren't traced to by others
            if req_id not in traced_to:
                summary["leaf_requirements"].append(req_id)

            # Orphan requirements have no relationships at all
            if not traces and not verified and req_id not in traced_to:
                summary["traceability"]["isolated"] += 1
                summary["orphan_requirements"].append(req_id)

        return summary
